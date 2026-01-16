"""Excel 导出工具。

用途：
- 蓝海监测结果导出
- 视频批处理结果导出

约束：
- 文件写入必须 try/except
- 错误信息要尽可能可被 Worker/UI 捕获并展示
"""

from __future__ import annotations

import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from typing import Callable, Dict, List, Optional
import config
from api.taobao_utils import generate_taobao_search_url

logger = logging.getLogger(__name__)


def _safe_emit(emit_log: Optional[Callable[[str], None]], message: str) -> None:
    if not emit_log:
        return
    try:
        emit_log(str(message))
    except Exception:
        pass


def _safe_float_percent(value) -> float:
    """将各种毛利表示安全转换为 float（百分比）。

    支持：
    - 50 / 50.0
    - "50%" / "50%+" / " 50 % "
    - None / "" / 非法字符串 -> 0.0
    """
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace("%", "").replace("+", "")
        if not text:
            return 0.0
        try:
            return float(text)
        except Exception:
            return 0.0
    return 0.0


def export_blue_ocean_results(
    products: List[Dict],
    filename: str = None,
    emit_log: Optional[Callable[[str], None]] = None,
) -> str:
    """导出蓝海监测结果到 Excel。

    Args:
        products: 商品列表
        filename: 文件名（为空则自动生成）
        emit_log: 可选日志回调（通常传入 Worker 的 self.emit_log）
    """
    try:
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"blue_ocean_{timestamp}.xlsx"

        filepath = config.OUTPUT_DIR / filename
        _safe_emit(emit_log, f"开始导出蓝海结果：{filepath}")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "蓝海选品结果"

        # Define headers
        headers = [
            "商品标题",
            "TK详情页",
            "1688搜索",
            "主图",
            "7日增长%",
            "评价数",
            "价格USD",
            "预估毛利%",
            "热门视频"
        ]

        # Apply header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # Add data rows
        for product in products:
            # Generate 1688 search link from title
            taobao_url = generate_taobao_search_url(product.get('title', ''))

            row = [
                product.get('title', ''),
                product.get('tk_url', ''),
                taobao_url,
                product.get('main_image_url', ''),
                product.get('growth_rate', 0),
                product.get('review_count', 0),
                product.get('price', 0),
                product.get('profit_margin', 0),
                product.get('top_video_url', '')
            ]

            ws.append(row)

            # Apply cell styling
            row_num = ws.max_row
            for col_num, cell in enumerate(ws[row_num], 1):
                cell.border = thin_border

                # Highlight high profit margin (> 20%)
                if col_num == 8 and _safe_float_percent(product.get('profit_margin', 0)) > 20:
                    cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

                # Center align numeric columns
                if col_num >= 5:
                    cell.alignment = Alignment(horizontal="center")

        # Adjust column widths
        column_widths = [20, 25, 30, 25, 12, 10, 12, 12, 25]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_num)].width = width

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save file
        wb.save(filepath)
        _safe_emit(emit_log, f"已导出蓝海结果到：{filepath}")
        return str(filepath)
    except Exception as e:
        logger.error(f"导出蓝海结果失败：{e}")
        _safe_emit(emit_log, f"导出蓝海结果失败：{e}")
        raise


def export_video_processing_log(
    videos: List[Dict],
    filename: str = None,
    emit_log: Optional[Callable[[str], None]] = None,
) -> str:
    """导出视频处理结果到 Excel。"""
    try:
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"video_processing_{timestamp}.xlsx"

        filepath = config.OUTPUT_DIR / filename
        _safe_emit(emit_log, f"开始导出视频处理日志：{filepath}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Video Processing Log"

        headers = [
            "原始文件名",
            "处理后文件名",
            "处理状态",
            "原始时长(秒)",
            "处理时长(秒)",
            "处理时间",
            "备注"
        ]

        ws.append(headers)

        for video in videos:
            ws.append([
                video.get('input_filename', ''),
                video.get('output_filename', ''),
                video.get('status', ''),
                video.get('original_duration', 0),
                video.get('processed_duration', 0),
                video.get('process_time', ''),
                video.get('notes', '')
            ])

        # Adjust column widths
        for col_num, width in enumerate([25, 25, 12, 15, 15, 20, 30], 1):
            ws.column_dimensions[chr(64 + col_num)].width = width

        wb.save(filepath)
        _safe_emit(emit_log, f"已导出视频处理日志到：{filepath}")
        return str(filepath)
    except Exception as e:
        logger.error(f"导出视频处理日志失败：{e}")
        _safe_emit(emit_log, f"导出视频处理日志失败：{e}")
        raise
